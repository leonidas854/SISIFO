"""Las hojas de cálculo no pueden salir todas iguales.

Un Excel genérico se nota y da pereza abrirlo. Cada trabajo debe tener su
aspecto, y si el usuario impone un estilo, ese manda sobre cualquier elección
automática.
"""
import openpyxl
import pytest

from dockit.generadores import estilos, xlsx

GUION = {"tipo": "xlsx", "titulo": "Datos", "bloques": [
    {"clase": "tabla", "cabecera": ["Concepto", "Valor", "Peso"],
     "filas": [["alfa", "10", "1.5"], ["beta", "20", "2.5"],
               ["gamma", "30", "3.5"]],
     "leyenda": "Tabla 1. Medidas", "fuente": "Elaboración propia"}]}


def test_hay_varios_temas_y_son_distintos():
    assert len(estilos.TEMAS) >= 4, "cuatro variantes es el mínimo para no repetirse"
    cabeceras = {t.cabecera_fondo for t in estilos.TEMAS.values()}
    assert len(cabeceras) == len(estilos.TEMAS), "dos temas con el mismo color no son dos temas"


def test_los_temas_difieren_en_algo_mas_que_el_color():
    rasgos = {(t.bandas, t.cabecera_negrita, t.borde) for t in estilos.TEMAS.values()}
    assert len(rasgos) > 1, "si solo cambia el color, es el mismo Excel pintado"


def test_elige_tema_estable_para_el_mismo_trabajo():
    a = estilos.elegir("cadena-custodia")
    b = estilos.elegir("cadena-custodia")
    assert a.nombre == b.nombre, "el mismo trabajo no puede cambiar de aspecto"


def test_trabajos_distintos_tienden_a_temas_distintos():
    nombres = {estilos.elegir(f"trabajo-{i}").nombre for i in range(12)}
    assert len(nombres) >= 3, "doce trabajos no pueden salir todos iguales"


def test_el_estilo_pedido_manda_sobre_el_automatico():
    t = estilos.elegir("lo-que-sea", pedido="sobrio")
    assert t.nombre == "sobrio"


def test_estilo_desconocido_no_revienta():
    t = estilos.elegir("x", pedido="no-existe-este-tema")
    assert t.nombre in estilos.TEMAS


@pytest.mark.parametrize("nombre", list(estilos.TEMAS))
def test_cada_tema_produce_un_xlsx_valido(tmp_path, nombre):
    destino = tmp_path / f"{nombre}.xlsx"
    xlsx.generar(GUION, str(destino), {}, {}, {"estilo": nombre})
    wb = openpyxl.load_workbook(str(destino))
    h = wb[wb.sheetnames[0]]
    assert h.cell(1, 1).value == "Concepto"
    assert h.cell(2, 2).value == 10, "los números deben seguir siendo números"
    assert h.cell(1, 1).fill.fgColor.rgb not in (None, "00000000"), \
        "la cabecera tiene que ir pintada"


def test_dos_trabajos_distintos_se_ven_distintos(tmp_path):
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    xlsx.generar(GUION, str(a), {}, {}, {"estilo": "sobrio"})
    xlsx.generar(GUION, str(b), {}, {}, {"estilo": "calido"})
    ca = openpyxl.load_workbook(str(a))["Tabla 1"].cell(1, 1).fill.fgColor.rgb
    cb = openpyxl.load_workbook(str(b))["Tabla 1"].cell(1, 1).fill.fgColor.rgb
    assert ca != cb
