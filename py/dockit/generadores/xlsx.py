"""Vuelca las tablas del guion a una hoja de cálculo.

Sirve para revisar los datos aparte del documento y para que OnlyOffice o
Excel puedan graficarlos sin volver a teclearlos.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import estilos
from . import guion as G

ANCHO_MAX = 60


def _bordes(tema):
    """Traduce el trato de borde del tema a objetos de openpyxl."""
    if tema.borde == "ninguno":
        return None, None
    fino = Side(style="thin", color="D5DAE0")
    acento = Side(style="medium", color=tema.acento)
    if tema.borde == "solo-cabecera":
        return Border(bottom=acento), None
    return Border(bottom=acento), Border(left=fino, right=fino, bottom=fino)


def _pintar_cabecera(hoja, fila, textos, tema, borde_cab):
    relleno = PatternFill("solid", fgColor=tema.cabecera_fondo)
    for j, texto in enumerate(textos, 1):
        c = hoja.cell(fila, j, texto)
        c.font = Font(bold=tema.cabecera_negrita, color=tema.cabecera_texto,
                      name=tema.fuente, size=11)
        c.fill = relleno
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if borde_cab is not None:
            c.border = borde_cab
    hoja.row_dimensions[fila].height = 22


def _nombre_hoja(leyenda: str, n: int, usados: set[str]) -> str:
    """Excel no admite más de 31 caracteres ni ciertos signos en el nombre."""
    base = re.split(r"[.:]", leyenda or "")[0].strip() or f"Tabla {n}"
    base = re.sub(r"[\[\]\*/\\?:]", " ", base).strip()[:31] or f"Tabla {n}"
    nombre, i = base, 1
    while nombre in usados:
        i += 1
        sufijo = f" ({i})"
        nombre = base[:31 - len(sufijo)] + sufijo
    usados.add(nombre)
    return nombre


def _ajustar(hoja) -> None:
    for col in hoja.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None),
                    default=8)
        hoja.column_dimensions[get_column_letter(col[0].column)].width = \
            min(ancho + 3, ANCHO_MAX)


def generar(guion: dict, destino: str, bibliografia: dict[str, str],
            en_texto: dict[str, str], formato: dict | None = None) -> dict:
    G.validar(guion, set(bibliografia) if bibliografia else None)

    formato = formato or {}
    # el estilo del BRIEF manda; si no hay, se reparte de forma estable por
    # el nombre del trabajo, para que dos trabajos no salgan idénticos
    tema = estilos.elegir(formato.get("_trabajo") or Path(destino).stem,
                          formato.get("estilo"))
    borde_cab, borde_cel = _bordes(tema)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    usados: set[str] = set()
    n = 0

    for b in guion["bloques"]:
        if b["clase"] != "tabla":
            continue
        n += 1
        hoja = wb.create_sheet(_nombre_hoja(b.get("leyenda", ""), n, usados))
        fila = 1
        if b.get("cabecera"):
            _pintar_cabecera(hoja, fila, b["cabecera"], tema, borde_cab)
            hoja.freeze_panes = "A2"
            fila += 1

        banda = PatternFill("solid", fgColor=tema.banda_fondo) \
            if tema.bandas and tema.banda_fondo else None
        for i, f in enumerate(b["filas"]):
            for j, valor in enumerate(f, 1):
                c = hoja.cell(fila, j, _numero_si_puede(valor))
                c.font = Font(name=tema.fuente, size=11, color=tema.texto)
                if banda is not None and i % 2 == 1:
                    c.fill = banda
                if borde_cel is not None:
                    c.border = borde_cel
                if isinstance(c.value, float):
                    c.number_format = "#,##0.00"
                elif isinstance(c.value, int):
                    c.number_format = "#,##0"
            fila += 1

        if b.get("fuente"):
            c = hoja.cell(fila + 1, 1, f"Fuente: {b['fuente']}")
            c.font = Font(italic=True, size=9, color=tema.acento, name=tema.fuente)
        _ajustar(hoja)

    if bibliografia:
        hoja = wb.create_sheet(_nombre_hoja("Referencias", n + 1, usados))
        _pintar_cabecera(hoja, 1, ["Clave", "Referencia (APA 7)"], tema, borde_cab)
        for i, (clave, entrada) in enumerate(sorted(bibliografia.items()), 2):
            hoja.cell(i, 1, clave)
            hoja.cell(i, 2, entrada).alignment = Alignment(wrap_text=True)
        hoja.column_dimensions["B"].width = ANCHO_MAX

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")

    Path(destino).parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return {"ruta": destino, "unidades": len(wb.sheetnames), "estilo": tema.nombre}


def _numero_si_puede(v):
    """Un número guardado como texto no se puede graficar ni sumar."""
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", ".")
    try:
        return int(s) if s.isdigit() or (s.startswith("-") and s[1:].isdigit()) \
            else float(s)
    except ValueError:
        return v
